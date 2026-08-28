"""
psers_data_pipeline.py — single source of truth for reading PSERS portfolio
data out of the source Excel workbook.

Every js/slide-data/*.data.js file that shows fund-level or portfolio-wide
numbers should ultimately trace back to a function in this module, not to
hand-typed values. If the workbook changes, re-run the relevant
tools/export_*.py script(s) and re-commit the regenerated data file(s) —
nothing here is meant to run automatically or at page-load time (this deck
has no build step and runs offline via file://), it's a manual/offline
generator you re-run when the source Excel changes.

SOURCE OF TRUTH
----------------
source-data/PSERS_Total Private Market Portfolio_Final.xlsx, sheet
"Clean Data" (NOT "Raw Data" — Raw Data is the pre-cleaning export; Clean
Data is what's actually been through Tangible's Include/Exclude review and
category mapping). Header labels live on ROW 2, not row 1 — row 1 is blank.
Data starts on row 3.

RULES THIS MODULE ENFORCES (per the user, 2026-08-27 — do not relitigate
these per-export, fix them here once):

1. Only "Include" rows count. The "Include / Exclude" column (Clean Data
   column Z) marks 363 of 744 total rows as "Include" as of the workbook
   checked in on 2026-08-26 — everything else (data errors, terminated
   positions, duplicates, whatever Tangible's review excluded) is dropped.
2. Every dollar figure uses the FIRST-listed (USD) column, never the
   "(Local)" one. Clean Data has both: Commitment / Paid in / Distributions
   / NAV / Total Value are the USD figures (already FX-converted using the
   "FX RATES — AS AT ..." table off to the side of this same sheet); their
   "(Local)" counterparts (Commitment (Local), Total Paid in (Local), Total
   Distribution (Local), Latest Valuation (Local), Unfunded Commitment
   (Local)) are the same positions in the manager's local currency and are
   NEVER used for anything in this deck.
3. "Called" (as this deck's slides label it) is the Excel column "Paid in"
   — same figure, different name. There is no separate "Called" column.
4. Ratios are ALWAYS derived from summed/individual dollar figures, never
   read from a pre-computed ratio column (Clean Data has none anyway) and
   never averaged across rows — sum the dollars first, then divide:
     DPI  = Distributions / Paid in
     RVPI = NAV / Paid in
     TVPI = (NAV + Distributions) / Paid in   (== Total Value / Paid in,
            since Clean Data's own "Total Value" column is already
            Distributions + NAV — spot-checked, not assumed)
5. "Revised Strategy (TANGIBLE)" is the strategy/category column used for
   grouping and display — NOT any of the four "Asset Class Tier N
   (Investment)" columns (those are a different classification, used
   elsewhere in this deck for the Strategy Deep Dive slide's tier1 filter).
   Clean Data's raw strategy values don't always match this deck's display
   labels — see STRATEGY_DISPLAY_LABELS below.

Everything below operates on USD-millions, 2-decimal-rounded, formatted the
same way the deck's existing hand-typed data files were: dollar figures
drop trailing zeros ("300" not "300.00", "165.33" stays as-is); ratios
always show exactly 2 decimals plus an "x" suffix ("0.00x", "1.03x").
"""

import json
import openpyxl

SOURCE_XLSX = "source-data/PSERS_Total Private Market Portfolio_Final.xlsx"
SHEET_NAME = "Clean Data"
HEADER_ROW = 2
DATA_START_ROW = 3

# Clean Data column layout (0-based index into a min_col=1,max_col=28 row
# tuple — i.e. index 0 is column A). Re-derived by reading the sheet's own
# header row, not hand-counted — see _verify_columns() below, which is run
# every time load_included_rows() is called so a reshuffled workbook fails
# loudly instead of silently mis-mapping a column.
#
# 2026-08-28 workbook added:
#   E  Strategy for Fund Level Metrics (splits Real Estate & Infra)
#   T  Unfunded Commitment ($) (USD — no FX conversion needed)
# and shifted later columns right. Dollar headers are now "Name ($)".
COL = {
    "investment": 2,           # C: Investment Name
    "strategy_raw": 3,         # D: Revised Strategy (TANGIBLE)
    "strategy_fund": 4,        # E: Strategy for Fund Level Metrics
    "portfolio": 5,            # F: Portfolio Name
    "asset": 6,                # G: Asset Name
    "manager": 7,              # H: Manager Name
    "vehicle": 8,              # I: Vehicle Type
    "tier1": 9,                # J: Asset Class Tier 1 (Investment)
    "tier2": 10,               # K: Asset Class Tier 2 (Investment)
    "tier3": 11,               # L: Asset Class Tier 3 (Investment)
    "tier4": 12,               # M: Asset Class Tier 4 (Investment)
    "vintage": 13,             # N: Vintage
    "commitment": 14,          # O: Commitment ($)
    "paid_in": 15,             # P: Paid in ($) — this deck's "Called"
    "distributions": 16,       # Q: Distributions ($)
    "nav": 17,                 # R: NAV ($)
    "total_value": 18,         # S: Total Value ($)
    "unfunded": 19,            # T: Unfunded Commitment ($)
    "valuation_date": 20,      # U: Latest Valuation Date
    "commitment_local": 21,    # V: Commitment (Local) — UNUSED
    "paid_in_local": 22,       # W: Total Paid in (Local) — UNUSED
    "distributions_local": 23, # X: Total Distribution (Local) — UNUSED
    "nav_local": 24,           # Y: Latest Valuation (Local) — UNUSED
    "unfunded_local": 25,      # Z: Unfunded Commitment (Local) — UNUSED
    "local_currency": 26,      # AA: Local Currency
    "include_exclude": 27,     # AB: Include / Exclude
}
DATA_MAX_COL = 28  # AB

EXPECTED_HEADERS = {
    "investment": "Investment Name",
    "strategy_raw": "Revised Strategy (TANGIBLE)",
    "strategy_fund": "Strategy for Fund Level Metrics",
    "manager": "Manager Name",
    "vintage": "Vintage",
    "commitment": "Commitment ($)",
    "paid_in": "Paid in ($)",
    "distributions": "Distributions ($)",
    "nav": "NAV ($)",
    "total_value": "Total Value ($)",
    "unfunded": "Unfunded Commitment ($)",
    "include_exclude": "Include / Exclude",
}

# Clean Data's raw "Revised Strategy (TANGIBLE)" values, mapped to the
# label this deck actually displays. Only "VC & Growth" differs — confirmed
# against every other slide that already shows a strategy label (Strategy
# Mix, Manager Concentration, the old Portfolio Holdings mock) all use
# "Growth & Venture", never the raw Excel string. Add new raw values here
# (not in each export script) if the workbook ever introduces one — a
# missing mapping raises, on purpose, rather than silently showing the raw
# Excel string on a slide.
STRATEGY_DISPLAY_LABELS = {
    "Real Estate & Infra": "Real Estate & Infra",
    "Private Credit": "Private Credit",
    "Private Equity": "Private Equity",
    "VC & Growth": "Growth & Venture",
}

# Column E "Strategy for Fund Level Metrics" — same four families as
# Revised Strategy, but Real Estate & Infra is already split.
FUND_LEVEL_STRATEGY_LABELS = {
    "Private Equity": "Private Equity",
    "Private Credit": "Private Credit",
    "VC & Growth": "Growth & Venture",
    "Real Estate": "Real Estate",
    "Infrastructure": "Infrastructure",
}


def _verify_columns(header_row_values):
    """Fail loudly if the sheet's actual header text doesn't match what
    COL above assumes — a reshuffled/edited workbook should break the
    export, not silently mis-map a column."""
    problems = []
    for key, expected in EXPECTED_HEADERS.items():
        idx = COL[key]
        actual = header_row_values[idx] if idx < len(header_row_values) else None
        if actual != expected:
            problems.append(f"  COL[{key!r}]={idx} expected header {expected!r}, found {actual!r}")
    if problems:
        raise RuntimeError(
            "Clean Data header layout has changed — update COL in "
            "psers_data_pipeline.py before trusting any export:\n" + "\n".join(problems)
        )


def load_fx_rates(path=SOURCE_XLSX):
    """Read the "FX RATES — AS AT ..." side-table that lives on the Clean
    Data sheet itself (to the right of the main data, starting around
    column AE) and return {currency: USD-per-1-unit rate}. There is no USD
    "Unfunded Commitment" column anywhere in this workbook (Raw Data or
    Clean Data) — only "Unfunded Commitment (Local)" — so this table is the
    only way to get that figure in USD (Round 71 used the same table by
    hand for the 40 non-USD CFG.rows funds it fixed; this function
    generalizes that to every row, every run, rather than a one-off patch).

    Located by searching column AG (index 32, 0-based) for the literal cell
    text "Currency" rather than a hardcoded row number — the table's exact
    row position isn't load-bearing, only its column and the "Currency" /
    "USD per 1 unit" header immediately above the rate rows. Reads rows
    until it hits a blank currency-name cell. Fails loudly (not silently)
    if the header text or column has moved, same philosophy as
    _verify_columns() above.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    fx_col = 32  # 0-based index of column AG (name); AH (index 33) holds the rate
    all_rows = list(ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=34, values_only=True))
    header_row_idx = None
    for i, r in enumerate(all_rows):
        if len(r) > fx_col and r[fx_col] == "Currency":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(
            "Could not find the FX RATES table's 'Currency' header in "
            "column AG of the Clean Data sheet — the table may have moved; "
            "update load_fx_rates() in psers_data_pipeline.py before "
            "trusting any unfunded/commitment_revised figure."
        )
    rates = {}
    for r in all_rows[header_row_idx + 1:]:
        name = r[fx_col] if len(r) > fx_col else None
        rate = r[fx_col + 1] if len(r) > fx_col + 1 else None
        if name is None:
            break
        if not isinstance(rate, (int, float)):
            raise RuntimeError(
                f"FX RATES table row for {name!r} has a non-numeric rate "
                f"{rate!r} — update load_fx_rates() before trusting this."
            )
        rates[name] = rate
    if "USD" not in rates:
        raise RuntimeError("FX RATES table was read but has no 'USD' entry — check load_fx_rates().")
    return rates


def load_included_rows(path=SOURCE_XLSX):
    """Read Clean Data, keep only Include rows, return a list of dicts with
    normalized field names and USD-only dollar figures (still in raw
    dollars, not millions — callers convert/format for their own slide)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(min_row=HEADER_ROW, min_col=1, max_col=DATA_MAX_COL, values_only=True))
    header_row = all_rows[0]
    _verify_columns(header_row)

    out = []
    for r in all_rows[DATA_START_ROW - HEADER_ROW:]:
        if r[COL["investment"]] is None:
            continue  # blank spacer row
        if r[COL["include_exclude"]] != "Include":
            continue

        strategy_raw = r[COL["strategy_raw"]]
        if strategy_raw not in STRATEGY_DISPLAY_LABELS:
            raise RuntimeError(
                f"Unmapped strategy value {strategy_raw!r} on row for "
                f"{r[COL['investment']]!r} — add it to STRATEGY_DISPLAY_LABELS "
                f"in psers_data_pipeline.py (don't guess a label per-export)."
            )
        strategy_fund_raw = r[COL["strategy_fund"]]
        if strategy_fund_raw not in FUND_LEVEL_STRATEGY_LABELS:
            raise RuntimeError(
                f"Unmapped fund-level strategy {strategy_fund_raw!r} on row for "
                f"{r[COL['investment']]!r} — add it to FUND_LEVEL_STRATEGY_LABELS "
                f"in psers_data_pipeline.py (don't guess a label per-export)."
            )

        commitment = r[COL["commitment"]] or 0
        paid_in = r[COL["paid_in"]] or 0
        distributions = r[COL["distributions"]] or 0
        nav = r[COL["nav"]] or 0
        total_value = r[COL["total_value"]] or 0
        unfunded = r[COL["unfunded"]] or 0

        out.append({
            "investment": r[COL["investment"]],
            "strategy_raw": strategy_raw,
            "strategy": STRATEGY_DISPLAY_LABELS[strategy_raw],
            "strategy_fund": FUND_LEVEL_STRATEGY_LABELS[strategy_fund_raw],
            "portfolio": r[COL["portfolio"]],
            "asset": r[COL["asset"]],
            "manager": r[COL["manager"]],
            "vehicle": r[COL["vehicle"]],
            "tier1": r[COL["tier1"]],
            "tier2": r[COL["tier2"]],
            "tier3": r[COL["tier3"]],
            "tier4": r[COL["tier4"]],
            "vintage": r[COL["vintage"]],
            "commitment": commitment,
            "paid_in": paid_in,
            "distributions": distributions,
            "nav": nav,
            "total_value": total_value,
            "unfunded": unfunded,
            "local_currency": r[COL["local_currency"]],
        })
    return out


def count_rows(path=SOURCE_XLSX):
    """Scan Clean Data once and return {"total", "included", "excluded"} —
    row counts only, no dollar figures, so this is NOT the "extraction"
    step build_clean_records() is (Round 69's single-extraction-pass rule
    is about never re-deriving per-row financial data in more than one
    place; a row tally isn't that). Used by exporters that need to state
    how many positions were excluded (e.g. Key Considerations' footnote)
    without re-implementing the Include/Exclude scan themselves."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(min_row=HEADER_ROW, min_col=1, max_col=DATA_MAX_COL, values_only=True))
    total = included = excluded = 0
    for r in all_rows[DATA_START_ROW - HEADER_ROW:]:
        if r[COL["investment"]] is None:
            continue
        total += 1
        if r[COL["include_exclude"]] == "Include":
            included += 1
        elif r[COL["include_exclude"]] == "Exclude":
            excluded += 1
    return {"total": total, "included": included, "excluded": excluded}


def ratios(paid_in, distributions, nav):
    """DPI / RVPI / TVPI from raw dollar figures — see rule 4 in the module
    docstring. paid_in == 0 returns zeros rather than raising, since a
    caller aggregating across many rows shouldn't crash on one all-zero
    bucket; as of the 2026-08-26 workbook no Included row actually has
    paid_in == 0 (checked), so this is a safety net, not an expected path."""
    if not paid_in:
        return 0.0, 0.0, 0.0
    dpi = distributions / paid_in
    rvpi = nav / paid_in
    tvpi = (nav + distributions) / paid_in
    return dpi, rvpi, tvpi


def fmt_millions(v):
    """An already-in-millions number -> display string: 2dp, trailing
    zeros stripped ('300' not '300.00', '165.33' stays). Use this for a
    value that's already in millions (e.g. a clean_data.json record's
    *_m field, or a sum of several of those) — it does NOT divide by 1e6.
    Matches the formatting already used throughout this deck's hand-typed
    slide-data files."""
    v2 = round(v, 2)
    s = f"{v2:.2f}"
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    if s in ("", "-0"):
        s = "0"
    return s


def fmt_money_m(raw_dollars):
    """Raw USD (not yet in millions) -> display string, via fmt_millions().
    Only needed by code working directly off load_included_rows()'s raw
    dollar figures — an exporter working off already-cleaned records (the
    normal case, see "Data pipeline" in README.md) should call
    fmt_millions() on that record's *_m field instead, not re-divide by
    1e6 itself."""
    return fmt_millions(raw_dollars / 1e6)


def fmt_ratio(v):
    """Ratio -> 2dp + 'x', never stripped ('0.00x', '1.00x') — matches the
    deck's existing DPI/RVPI/TVPI formatting everywhere it appears."""
    return f"{v:.2f}x"


def js_string_literal(s):
    """Escape a Python string for embedding inside a single-quoted JS
    string literal in a generated .data.js file."""
    return str(s).replace("\\", "\\\\").replace("'", "\\'")


def unfunded_pct(unfunded, commitment):
    """Unfunded $ / Commitment $, as a 0–1 fraction.
    Capped at 1.0 (100%). Commitment of 0 (or any error) → 0."""
    try:
        c = float(commitment or 0)
        if c == 0:
            return 0.0
        p = float(unfunded or 0) / c
        if p != p:  # NaN
            return 0.0
        return round(min(1.0, p), 6)
    except (TypeError, ValueError, ZeroDivisionError):
        return 0.0


def to_millions_record(r):
    """Convert one raw-dollar row (a dict from load_included_rows()) into a
    JSON-friendly record: dollar figures as numeric millions (not the
    2dp-stripped display strings the .data.js exporters use — this is for
    reuse/analysis, not direct slide rendering), plus computed DPI/RVPI/TVPI.
    Used by dump_clean_data_json() below, and available to any future
    exporter that wants numeric millions instead of raw dollars or a
    display-formatted string."""
    dpi, rvpi, tvpi = ratios(r["paid_in"], r["distributions"], r["nav"])
    record = {
        "investment": r["investment"],
        "strategy": r["strategy"],
        "strategy_raw": r["strategy_raw"],
        "strategy_fund": r["strategy_fund"],
        "portfolio": r["portfolio"],
        "asset": r["asset"],
        "manager": r["manager"],
        "vehicle": r["vehicle"],
        "tier1": r["tier1"],
        "tier2": r["tier2"],
        "tier3": r["tier3"],
        "tier4": r["tier4"],
        "vintage": r["vintage"],
        # 6dp here (not the 2dp any display ever shows) is deliberate:
        # exporters round AGAIN to 2dp for display (fmt_millions/fmt_ratio),
        # and rounding twice at close precisions can shift the last displayed
        # digit vs. rounding the raw value once (double-rounding) — found
        # this for real in Round 69 (e.g. a DPI landing on .895 rounded here
        # to 4dp first came out .90x instead of the correct .89x). 6dp is far
        # enough above 2dp that this can't recur for figures at this scale;
        # it's headroom for the *display* step's rounding, not the record's
        # own precision.
        "commitment_m": round(r["commitment"] / 1e6, 6),
        "paid_in_m": round(r["paid_in"] / 1e6, 6),
        "distributions_m": round(r["distributions"] / 1e6, 6),
        "nav_m": round(r["nav"] / 1e6, 6),
        "total_value_m": round(r["total_value"] / 1e6, 6),
        "dpi": round(dpi, 6),
        "rvpi": round(rvpi, 6),
        "tvpi": round(tvpi, 6),
        # unfunded_m / commitment_revised_m / unfunded_pct / funded_pct:
        # added for the CFG.rows migration (tools/export_cfg_rows.py) but
        # kept here, not exporter-local, since they're general portfolio
        # metrics (any future exporter needing "% unfunded" should reuse
        # these, not re-derive from unfunded/FX itself -- same reasoning as
        # every other field in this record).
        # unfunded_pct = Unfunded $ / Commitment $, capped at 100%;
        # zero commitment (or any error) falls back to 0.
        # funded_pct stays paid / (paid + unfunded) and is independent.
        "unfunded_m": round(r["unfunded"] / 1e6, 6),
        "commitment_revised_m": round(r["paid_in"] / 1e6 + r["unfunded"] / 1e6, 6),
        "unfunded_pct": unfunded_pct(r["unfunded"], r["commitment"]),
    }
    paid_plus_unfunded = r["paid_in"] + r["unfunded"]
    record["funded_pct"] = (
        round(r["paid_in"] / paid_plus_unfunded, 6) if paid_plus_unfunded else 0.0
    )
    return record


def build_clean_records(included_rows):
    """Turn the raw rows from load_included_rows() into the cleaned,
    computed form (millions + DPI/RVPI/TVPI) ONCE. This is the single
    "extraction" step for the whole pipeline — tools/export.py calls this
    exactly once per run and hands the SAME list to dump_clean_data_json()
    AND every registered slide exporter, so no exporter re-derives
    millions/ratios from raw dollars itself (that used to happen twice —
    once here, once again inside export_portfolio_holdings.py — fixed in
    Round 69). A slide exporter's build_rows() should only format/sort/
    aggregate these records, never recompute a ratio or re-divide by 1e6."""
    return [to_millions_record(r) for r in included_rows]


def dump_clean_data_json(records, path):
    """Write an already-built list of clean records (from
    build_clean_records()) as one general-purpose JSON file — not shaped
    for any particular slide. This is the "do whatever we want with it
    later" artifact: a future exporter (in Python OR plain JS, e.g. if a
    slide ever needs to read data without a Python step) can load this
    instead of re-parsing the workbook, and it's a fast way to spot-check
    the full extracted dataset by eye. Regenerated every run by
    tools/export.py — like the .data.js files, don't hand-edit it."""
    payload = {
        "generated_from": SOURCE_XLSX,
        "sheet": SHEET_NAME,
        "count": len(records),
        "records": records,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return path
